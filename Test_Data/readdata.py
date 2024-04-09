import openpyxl
wk = openpyxl.load_workbook("C:/Users/Vinay/Desktop/Vinay/TestData.xlsx")

def fetch_no_of_rows(sheetname):
    sh = wk[sheetname]
    return sh.max_row

def fetch_cell_data(sheetname, row, col):
    sh = wk[sheetname]
    cell = sh.cell(int(row),int(col))
    return cell.value

print(fetch_cell_data("Sheet1",1,2))